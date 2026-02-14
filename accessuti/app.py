
from flask import Flask, render_template, request, redirect, session, url_for, flash
import csv
import os
import json
import heapq
from openpyxl import load_workbook
from datetime import datetime, date

from werkzeug.security import generate_password_hash, check_password_hash

from estructuras.bst import ArbolUsuarios
from estructuras.pila import Pila


# ================= APP =================
app = Flask(__name__)
app.secret_key = "accessuti-final"

BASE_DIR = app.root_path
DATA_DIR = os.path.join(BASE_DIR, "data")

USUARIOS_SISTEMA = os.path.join(DATA_DIR, "usuarios_sistema.csv")
USUARIOS_RED = os.path.join(DATA_DIR, "usuarios_red.csv")
AUDITORIA_CSV = os.path.join(DATA_DIR, "auditoria.csv")
ORG_XLSX = os.path.join(DATA_DIR, "Organizacion INIA.xlsx")


# ================= ESTRUCTURAS =================
arbol_sistema = ArbolUsuarios()
arbol_red = ArbolUsuarios()
auditoria = Pila()


# ================= CATÁLOGOS =================
CAT_SEDES = []
CAT_CONTRATOS = ["CAS", "Nombrado", "Tercero", "Practicante", "Locación de servicios", "Otro"]
CAT_NIVEL_RED = ["NORMAL", "ACCESO REDES SOCIALES", "ACCESO REMOTO", "LIBRE"]

MAP_SEDE_DIR = {}   # sede -> dependencias
MAP_DIR_SUB = {}    # dependencia -> subdependencias


# ================= ORGANIZACIÓN (EXCEL) =================
def limpiar_txt(x):
    return (str(x).replace("  ", " ").strip()) if x is not None else ""


def cargar_organizacion():
    """
    Lee Organizacion INIA.xlsx y arma:
      - CAT_SEDES
      - MAP_SEDE_DIR: sede -> [dependencias]
      - MAP_DIR_SUB: dependencia -> [subdependencias]
    """
    global CAT_SEDES, MAP_SEDE_DIR, MAP_DIR_SUB

    CAT_SEDES = []
    MAP_SEDE_DIR = {}
    MAP_DIR_SUB = {}

    if not os.path.exists(ORG_XLSX):
        print("⚠️ No se encontró Organizacion INIA.xlsx en:", ORG_XLSX)
        return

    wb = load_workbook(ORG_XLSX)
    ws = wb.active

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return

    header = [limpiar_txt(x).upper() for x in filas[0]]

    try:
        i_sede = header.index("SEDE")
        i_dir = header.index("DIRECCION")
        i_sub = header.index("SUBDIRECCION")
    except ValueError:
        print("⚠️ Encabezados esperados no encontrados. Detectado:", header)
        return

    for row in filas[1:]:
        sede = limpiar_txt(row[i_sede])
        dire = limpiar_txt(row[i_dir])
        sub = limpiar_txt(row[i_sub])

        if not sede:
            sede = "OFICINA CENTRAL"

        # Siempre incluir sede (incluye EEA aunque dirección sea "-")
        CAT_SEDES.append(sede)
        MAP_SEDE_DIR.setdefault(sede, set())

        if dire in ("", "-"):
            continue

        if sub.upper() in ("", "-", "NONE"):
            sub = ""

        MAP_SEDE_DIR[sede].add(dire)
        MAP_DIR_SUB.setdefault(dire, set())
        if sub:
            MAP_DIR_SUB[dire].add(sub)

    CAT_SEDES = sorted(list(set(CAT_SEDES)))
    for k in list(MAP_SEDE_DIR.keys()):
        MAP_SEDE_DIR[k] = sorted(list(MAP_SEDE_DIR[k]))
    for k in list(MAP_DIR_SUB.keys()):
        MAP_DIR_SUB[k] = sorted(list(MAP_DIR_SUB[k]))

    print("✔ SEDES CARGADAS:", len(CAT_SEDES))


# ================= FECHAS / ALERTAS =================
def hoy():
    return date.today()


def dias_restantes(f):
    if not f:
        return None
    try:
        return (date.fromisoformat(f) - hoy()).days
    except Exception:
        return None


# ================= AUDITORIA (PILA + CSV) =================
def asegurar_csv_con_header(path, headers):
    if (not os.path.exists(path)) or os.path.getsize(path) == 0:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()


def registrar(user, accion, detalle=""):
    item = {
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "usuario": user,
        "accion": accion,
        "detalle": detalle
    }
    auditoria.push(item)

    headers = ["fecha", "usuario", "accion", "detalle"]
    asegurar_csv_con_header(AUDITORIA_CSV, headers)
    with open(AUDITORIA_CSV, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writerow(item)


def auditoria_reciente(limit=12):
    # Pila: top -> bottom
    items = auditoria.to_list()
    return items[:limit]


# ================= AUTH HELPERS =================
def require_login():
    return session.get("user")


def get_user_record(username):
    return arbol_sistema.buscar(username) if username else None


def is_admin_user(username):
    us = get_user_record(username) or {}
    return (us.get("rol") or "").strip().upper() == "ADMIN"


def login_required(fn):
    def wrapper(*args, **kwargs):
        if not require_login():
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    wrapper.__name__ = fn.__name__
    return wrapper


def admin_required(fn):
    def wrapper(*args, **kwargs):
        u = require_login()
        if not u:
            return redirect(url_for("login"))
        if not is_admin_user(u):
            flash("Acceso restringido: solo ADMIN.", "danger")
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    wrapper.__name__ = fn.__name__
    return wrapper


# ================= CSV: USUARIOS RED =================
def headers_usuario_red():
    return [
        "usuario_red", "nombres", "tipo_contrato", "sede", "dependencia", "subdependencia",
        "estado", "fecha_inicio", "fecha_fin",
        "nivel_red", "red_inicio", "red_fin",
        "tiene_vpn", "vpn_inicio", "vpn_fin"
    ]


def append_usuario_red_csv(row):
    headers = headers_usuario_red()
    for h in headers:
        row.setdefault(h, "")

    asegurar_csv_con_header(USUARIOS_RED, headers)

    with open(USUARIOS_RED, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writerow({h: row.get(h, "") for h in headers})


def update_usuario_red_csv(usuario_red, nuevos_campos: dict) -> bool:
    if not os.path.exists(USUARIOS_RED):
        return False

    headers = headers_usuario_red()
    filas = []
    encontrado = False

    with open(USUARIOS_RED, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            if (r.get("usuario_red") or "").strip() == usuario_red:
                for k, v in nuevos_campos.items():
                    if k in headers:
                        r[k] = v
                encontrado = True
            filas.append(r)

    if not encontrado:
        return False

    with open(USUARIOS_RED, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for r in filas:
            writer.writerow({h: r.get(h, "") for h in headers})

    return True


# ================= CSV: USUARIOS SISTEMA =================
def headers_usuario_sistema():
    return ["username", "password_hash", "rol", "estado", "creado_en"]


def append_usuario_sistema_csv(row):
    headers = headers_usuario_sistema()
    for h in headers:
        row.setdefault(h, "")

    asegurar_csv_con_header(USUARIOS_SISTEMA, headers)

    with open(USUARIOS_SISTEMA, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writerow({h: row.get(h, "") for h in headers})


# ================= CARGAS =================
def cargar_sistema():
    """
    - Crea usuarios_sistema.csv si no existe
    - Soporta CSV viejo con columna 'password' (migra a password_hash)
    """
    global arbol_sistema
    arbol_sistema = ArbolUsuarios()

    os.makedirs(DATA_DIR, exist_ok=True)

    if not os.path.exists(USUARIOS_SISTEMA):
        asegurar_csv_con_header(USUARIOS_SISTEMA, headers_usuario_sistema())
        append_usuario_sistema_csv({
            "username": "admin",
            "password_hash": generate_password_hash("admin123"),
            "rol": "ADMIN",
            "estado": "ACTIVO",
            "creado_en": datetime.now().strftime("%Y-%m-%d %H:%M")
        })
        print("✔ usuarios_sistema.csv creado con ADMIN admin/admin123")

    with open(USUARIOS_SISTEMA, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = [c.strip() for c in (reader.fieldnames or [])]

        es_esquema_viejo = ("password" in fieldnames) and ("password_hash" not in fieldnames)

        filas_norm = []
        for r in reader:
            u = (r.get("username") or "").strip()
            if not u:
                continue

            rol = (r.get("rol") or "USER").strip().upper()
            estado = (r.get("estado") or "ACTIVO").strip().upper()

            if es_esquema_viejo:
                pw_plain = (r.get("password") or "").strip()
                pw_hash = generate_password_hash(pw_plain) if pw_plain else ""
                creado = datetime.now().strftime("%Y-%m-%d %H:%M")
            else:
                pw_hash = (r.get("password_hash") or "").strip()
                creado = (r.get("creado_en") or "").strip()

            registro = {
                "username": u,
                "password_hash": pw_hash,
                "rol": rol,
                "estado": estado,
                "creado_en": creado,
            }

            filas_norm.append(registro)
            arbol_sistema.insertar(u, registro)

    if es_esquema_viejo:
        asegurar_csv_con_header(USUARIOS_SISTEMA, headers_usuario_sistema())
        with open(USUARIOS_SISTEMA, "w", newline="", encoding="utf-8") as wf:
            w = csv.DictWriter(wf, fieldnames=headers_usuario_sistema())
            w.writeheader()
            for reg in filas_norm:
                w.writerow({
                    "username": reg["username"],
                    "password_hash": reg["password_hash"],
                    "rol": reg["rol"],
                    "estado": reg["estado"],
                    "creado_en": reg["creado_en"],
                })
        print("✔ Migración realizada: usuarios_sistema.csv ahora usa password_hash")


def cargar_red():
    global arbol_red
    arbol_red = ArbolUsuarios()

    if not os.path.exists(USUARIOS_RED):
        return

    with open(USUARIOS_RED, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            u = (r.get("usuario_red") or "").strip()
            if not u:
                continue

            r["usuario_red"] = u
            r["nombres"] = (r.get("nombres") or "").strip()
            r["tipo_contrato"] = (r.get("tipo_contrato") or "").strip()
            r["sede"] = (r.get("sede") or "").strip()
            r["dependencia"] = (r.get("dependencia") or "").strip()
            r["subdependencia"] = (r.get("subdependencia") or "").strip()

            r["estado"] = (r.get("estado") or "INACTIVO").strip().upper()
            r["nivel_red"] = (r.get("nivel_red") or "NORMAL").strip().upper()
            r["tiene_vpn"] = (r.get("tiene_vpn") or "NO").strip().upper()

            r["fecha_inicio"] = (r.get("fecha_inicio") or "").strip()
            r["fecha_fin"] = (r.get("fecha_fin") or "").strip()
            r["red_inicio"] = (r.get("red_inicio") or "").strip()
            r["red_fin"] = (r.get("red_fin") or "").strip()
            r["vpn_inicio"] = (r.get("vpn_inicio") or "").strip()
            r["vpn_fin"] = (r.get("vpn_fin") or "").strip()

            arbol_red.insertar(u, r)


# ================= DASHBOARD: SERIES (INTERACTIVO) =================
def agg_counts_heap(lista, field, top_n=12):
    """
    Cuenta ocurrencias de 'field' y retorna Top-N usando heap.
    Retorna: { total, items:[{label,count,pct}], otros_count }
    """
    total = len(lista)
    counts = {}

    for x in lista:
        v = (x.get(field) or "").strip()
        if v == "":
            v = "—"
        counts[v] = counts.get(v, 0) + 1

    # heapq.nlargest devuelve Top N por count (sin ordenar todo si no hace falta)
    top = heapq.nlargest(top_n, counts.items(), key=lambda kv: kv[1])

    items = []
    top_sum = 0
    for label, count in top:
        top_sum += count
        pct = int(round((count * 100) / total)) if total else 0
        items.append({"label": label, "count": count, "pct": pct})

    otros = max(0, total - top_sum)

    return {
        "total": total,
        "items": items,
        "otros_count": otros
    }


def build_dashboard_series_topn(usuarios, dim, top_n=12):
    """
    dim: SEDE | DEPENDENCIA | SUBDEPENDENCIA | NIVEL_RED | VPN | ESTADO | CONTRATO
    """
    dim = (dim or "SEDE").upper()
    top_n = int(top_n or 12)

    field_map = {
        "SEDE": "sede",
        "DEPENDENCIA": "dependencia",
        "SUBDEPENDENCIA": "subdependencia",
        "NIVEL_RED": "nivel_red",
        "VPN": "tiene_vpn",
        "ESTADO": "estado",
        "CONTRATO": "tipo_contrato",
    }

    field = field_map.get(dim, "sede")
    return agg_counts_heap(usuarios, field, top_n=top_n)

def agg_counts(lista, field):
    """
    Retorna lista [{label,count,pct}] ordenado desc para un campo (field)
    """
    total = len(lista)
    counts = {}
    for x in lista:
        v = (x.get(field) or "—").strip()
        if v == "":
            v = "—"
        counts[v] = counts.get(v, 0) + 1

    out = []
    for label, count in counts.items():
        pct = int(round((count * 100) / total)) if total else 0
        out.append({"label": label, "count": count, "pct": pct})

    out.sort(key=lambda a: a["count"], reverse=True)
    return out


def build_dashboard_series(usuarios):
    """
    Arma todas las series que el dashboard puede mostrar
    """
    return {
        "SEDE": agg_counts(usuarios, "sede"),
        "DEPENDENCIA": agg_counts(usuarios, "dependencia"),
        "SUBDEPENDENCIA": agg_counts(usuarios, "subdependencia"),
        "ESTADO": agg_counts(usuarios, "estado"),
        "VPN": agg_counts(usuarios, "tiene_vpn"),
        "NIVEL_RED": agg_counts(usuarios, "nivel_red"),
        "CONTRATO": agg_counts(usuarios, "tipo_contrato"),
    }


# ================= INIT =================
cargar_organizacion()
cargar_sistema()
cargar_red()


# ================= ROUTES =================
@app.get("/")
def home():
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if require_login():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = (request.form.get("password") or "").strip()

        user = arbol_sistema.buscar(u)
        if not user:
            return render_template("login.html", error="Usuario no existe")

        if (user.get("estado") or "").strip().upper() != "ACTIVO":
            return render_template("login.html", error="Usuario inactivo")

        ph = (user.get("password_hash") or "").strip()
        if not ph or (not check_password_hash(ph, p)):
            return render_template("login.html", error="Contraseña incorrecta")

        session["user"] = u
        registrar(u, "LOGIN_OK")
        return redirect(url_for("dashboard"))

    return render_template("login.html", error=None)


@app.route("/register", methods=["GET", "POST"])
def register():
    # registro libre (si quieres, lo hacemos solo ADMIN)
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = (request.form.get("password") or "").strip()
        p2 = (request.form.get("password2") or "").strip()

        if len(u) < 3:
            return render_template("register.html", error="El usuario debe tener mínimo 3 caracteres.")
        if p != p2:
            return render_template("register.html", error="Las contraseñas no coinciden.")
        if len(p) < 6:
            return render_template("register.html", error="La contraseña debe tener mínimo 6 caracteres.")
        if arbol_sistema.buscar(u) is not None:
            return render_template("register.html", error="Ese usuario ya existe.")

        append_usuario_sistema_csv({
            "username": u,
            "password_hash": generate_password_hash(p),
            "rol": "USER",
            "estado": "ACTIVO",
            "creado_en": datetime.now().strftime("%Y-%m-%d %H:%M")
        })
        cargar_sistema()
        registrar(u, "REGISTER_OK")
        flash("Registro exitoso. Ya puedes iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("register.html", error=None)


@app.get("/logout")
def logout():
    u = session.get("user")
    if u:
        registrar(u, "LOGOUT")
    session.clear()
    return redirect(url_for("login"))


@app.get("/dashboard")
@login_required
def dashboard():
    u = require_login()
    usuarios = arbol_red.inorder()

    total = len(usuarios)
    activos = sum(1 for x in usuarios if x.get("estado") == "ACTIVO")
    especiales = sum(1 for x in usuarios if (x.get("nivel_red") or "NORMAL") != "NORMAL")
    vpn = sum(1 for x in usuarios if (x.get("tiene_vpn") or "NO") == "SI")

    # Alertas por vencimiento (7 días)
    alert_items = []
    for x in usuarios:
        nivel = (x.get("nivel_red") or "NORMAL")
        if nivel != "NORMAL":
            d = dias_restantes(x.get("red_fin"))
            if d is not None and d <= 7:
                alert_items.append({
                    "usuario_red": x.get("usuario_red"),
                    "nombres": x.get("nombres"),
                    "dependencia": x.get("dependencia"),
                    "tipo": "ACCESO RED",
                    "fin": x.get("red_fin") or "—",
                    "dias": d,
                })

        if (x.get("tiene_vpn") or "NO") == "SI":
            d2 = dias_restantes(x.get("vpn_fin"))
            if d2 is not None and d2 <= 7:
                alert_items.append({
                    "usuario_red": x.get("usuario_red"),
                    "nombres": x.get("nombres"),
                    "dependencia": x.get("dependencia"),
                    "tipo": "VPN",
                    "fin": x.get("vpn_fin") or "—",
                    "dias": d2,
                })

    alert_items.sort(key=lambda a: a["dias"])
    alertas = len(alert_items)

    series = build_dashboard_series(usuarios)

    return render_template(
        "dashboard.html",
        user=u,
        is_admin=is_admin_user(u),
        total=total,
        activos=activos,
        especiales=especiales,
        vpn=vpn,
        alertas=alertas,
        alert_items=alert_items[:10],
        audit_items=auditoria_reciente(12),
        series_json=json.dumps(series, ensure_ascii=False),
    )


@app.context_processor
def inject_globals():
    u = session.get("user")
    return {"is_admin": is_admin_user(u) if u else False}


@app.route("/admin/usuarios_red/nuevo", methods=["GET", "POST"])
@admin_required
def nuevo_usuario_red():
    admin = require_login()
    error = None

    if request.method == "POST":
        row = {k: (request.form.get(k, "") or "").strip() for k in headers_usuario_red()}
        row["estado"] = (row.get("estado") or "ACTIVO").upper()
        row["nivel_red"] = (row.get("nivel_red") or "NORMAL").upper()
        row["tiene_vpn"] = (row.get("tiene_vpn") or "NO").upper()

        if not row["usuario_red"] or not row["nombres"]:
            error = "Completa Usuario de red y Nombres."
        elif arbol_red.buscar(row["usuario_red"]) is not None:
            error = "Ese usuario de red ya existe."
        else:
            append_usuario_red_csv(row)
            cargar_red()
            registrar(admin, "REGISTRO_USUARIO_RED", row["usuario_red"])
            flash("Usuario de red registrado.", "success")
            return redirect(url_for("consulta_usuario_red", ver=row["usuario_red"]))

    return render_template(
        "usuario_red_form.html",
        user=admin,
        is_admin=True,
        error=error,
        sedes=CAT_SEDES,
        contratos=CAT_CONTRATOS,
        niveles_red=CAT_NIVEL_RED,
        map_sede_dir_json=json.dumps(MAP_SEDE_DIR, ensure_ascii=False),
        map_dir_sub_json=json.dumps(MAP_DIR_SUB, ensure_ascii=False),
    )


@app.get("/admin/usuarios_red/editar/<usuario_red>")
@admin_required
def editar_usuario_red_ui(usuario_red):
    admin = require_login()
    udata = arbol_red.buscar(usuario_red)
    if not udata:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("consulta_usuario_red"))

    return render_template(
        "usuario_red_edit.html",
        user=admin,
        is_admin=True,
        usuario=udata,
        error=None,
        sedes=CAT_SEDES,
        contratos=CAT_CONTRATOS,
        niveles_red=CAT_NIVEL_RED,
        map_sede_dir_json=json.dumps(MAP_SEDE_DIR, ensure_ascii=False),
        map_dir_sub_json=json.dumps(MAP_DIR_SUB, ensure_ascii=False),
    )


@app.post("/admin/usuarios_red/editar/<usuario_red>")
@admin_required
def editar_usuario_red_post(usuario_red):
    admin = require_login()
    udata = arbol_red.buscar(usuario_red)
    if not udata:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("consulta_usuario_red"))

    cambios = {k: (request.form.get(k, "") or "").strip() for k in headers_usuario_red()}
    cambios["estado"] = (cambios.get("estado") or "ACTIVO").upper()
    cambios["nivel_red"] = (cambios.get("nivel_red") or "NORMAL").upper()
    cambios["tiene_vpn"] = (cambios.get("tiene_vpn") or "NO").upper()

    cambios.pop("usuario_red", None)  # no se edita

    if not cambios.get("nombres"):
        return render_template(
            "usuario_red_edit.html",
            user=admin,
            is_admin=True,
            usuario=udata,
            error="Nombres no puede estar vacío.",
            sedes=CAT_SEDES,
            contratos=CAT_CONTRATOS,
            niveles_red=CAT_NIVEL_RED,
            map_sede_dir_json=json.dumps(MAP_SEDE_DIR, ensure_ascii=False),
            map_dir_sub_json=json.dumps(MAP_DIR_SUB, ensure_ascii=False),
        )

    ok = update_usuario_red_csv(usuario_red, cambios)
    if ok:
        cargar_red()
        registrar(admin, "UPDATE_USUARIO_RED", usuario_red)
        flash("Cambios guardados.", "success")
    else:
        flash("No se pudo actualizar (CSV).", "danger")

    next_url = request.args.get("next")
    if next_url:
        return redirect(next_url)
    return redirect(url_for("consulta_usuario_red", ver=usuario_red))


@app.get("/admin/usuarios_red/toggle/<usuario_red>")
@admin_required
def toggle_estado_usuario_red(usuario_red):
    admin = require_login()

    udata = arbol_red.buscar(usuario_red)
    if not udata:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("consulta_usuario_red"))

    estado_actual = (udata.get("estado") or "INACTIVO").upper()
    nuevo = "INACTIVO" if estado_actual == "ACTIVO" else "ACTIVO"

    ok = update_usuario_red_csv(usuario_red, {"estado": nuevo})
    if ok:
        cargar_red()
        registrar(admin, "TOGGLE_ESTADO_USUARIO_RED", f"{usuario_red}->{nuevo}")
        flash(f"Estado cambiado a {nuevo}.", "success")
    else:
        flash("No se pudo cambiar el estado.", "danger")

    next_url = request.args.get("next")
    if next_url:
        return redirect(next_url)
    return redirect(url_for("consulta_usuario_red", ver=usuario_red))


@app.get("/consulta_usuario_red")
@login_required
def consulta_usuario_red():
    u = require_login()

    q = (request.args.get("q") or "").strip().lower()
    f_sede = (request.args.get("sede") or "").strip()
    f_dep = (request.args.get("dependencia") or "").strip()
    f_sub = (request.args.get("subdependencia") or "").strip()
    f_estado = (request.args.get("estado") or "").strip().upper()
    f_nivel = (request.args.get("nivel_red") or "").strip().upper()
    f_vpn = (request.args.get("tiene_vpn") or "").strip().upper()

    usuarios = arbol_red.inorder()
    filtrados = []

    for x in usuarios:
        if q:
            t1 = (x.get("usuario_red") or "").lower()
            t2 = (x.get("nombres") or "").lower()
            t3 = (x.get("dependencia") or "").lower()
            if q not in t1 and q not in t2 and q not in t3:
                continue

        if f_sede and (x.get("sede") or "") != f_sede:
            continue
        if f_dep and (x.get("dependencia") or "") != f_dep:
            continue
        if f_sub and (x.get("subdependencia") or "") != f_sub:
            continue
        if f_estado and (x.get("estado") or "") != f_estado:
            continue
        if f_nivel and (x.get("nivel_red") or "") != f_nivel:
            continue
        if f_vpn and (x.get("tiene_vpn") or "") != f_vpn:
            continue

        filtrados.append(x)

    usuario_detalle = None
    ver = (request.args.get("ver") or "").strip()
    if ver:
        usuario_detalle = arbol_red.buscar(ver)

    return render_template(
        "consulta_usuario_red.html",
        user=u,
        is_admin=is_admin_user(u),
        q=q,
        filtros={
            "sede": f_sede,
            "dependencia": f_dep,
            "subdependencia": f_sub,
            "estado": f_estado,
            "nivel_red": f_nivel,
            "tiene_vpn": f_vpn,
        },
        resultados=filtrados,
        usuario=usuario_detalle,
        sedes=CAT_SEDES,
        contratos=CAT_CONTRATOS,
        niveles_red=CAT_NIVEL_RED,
        map_sede_dir_json=json.dumps(MAP_SEDE_DIR, ensure_ascii=False),
        map_dir_sub_json=json.dumps(MAP_DIR_SUB, ensure_ascii=False),
    )


if __name__ == "__main__":
    app.run(debug=True, port=5001)